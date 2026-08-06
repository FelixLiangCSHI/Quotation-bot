from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "Decision Tree"
OUTPUT_DIR = ROOT / "rules"
NORMALIZED_JSON = OUTPUT_DIR / "decision_tree_normalized_rules.json"
REVIEW_CSV = OUTPUT_DIR / "decision_tree_rules_needing_review.csv"

MAX_SCAN_COLUMNS = 120
PRODUCT_ID_RE = re.compile(r"^\d{6,8}$")
STEP_RE = re.compile(r"\bstep\s*([0-9]+[a-z]?)\b", re.IGNORECASE)

SKIP_SHEET_TOKENS = ("pli_doc_info", "notes")

REVIEW_COLUMNS = [
    "review_id",
    "rule_id",
    "rule_type",
    "current_review_status",
    "engine_status",
    "review_reason",
    "product_line",
    "product_id",
    "step_id",
    "option_group",
    "regions",
    "strength",
    "confidence",
    "message",
    "source_workbook",
    "source_sheet",
    "source_cell",
    "normalized_payload",
    "review_decision",
    "final_rule_type",
    "final_effect",
    "reviewer",
    "reviewed_at",
    "review_notes",
]

REGION_PATTERNS = {
    "canada": re.compile(r"\b(canada|canadian|ca)\b", re.IGNORECASE),
    "china": re.compile(r"\b(china|prc|cn)\b", re.IGNORECASE),
    "us": re.compile(r"\b(us|u\.s\.|usa|united states)\b", re.IGNORECASE),
}

RULE_KEYWORDS = (
    "only",
    "must",
    "need",
    "required",
    "requires",
    "choose",
    "pick",
    "select",
    "cannot",
    "can't",
    "not compatible",
    "not support",
    "not include",
    "exclude",
    "except",
    "china",
    "canada",
    "us",
    "warning",
    "confirm",
    "follow",
)


@dataclass(frozen=True)
class StepContext:
    step_id: str = ""
    option_group: str = ""


def main() -> None:
    try:
        import openpyxl
    except ImportError as error:
        raise SystemExit(
            "openpyxl is required to normalize decision tree workbooks. "
            "Install dependencies with: pip install -r requirements.txt"
        ) from error

    source_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else SOURCE_DIR
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    products: list[dict[str, Any]] = []
    rules: list[dict[str, Any]] = []
    workbook_summaries: list[dict[str, Any]] = []

    workbooks = sorted(source_dir.glob("*.xlsx"))
    if not workbooks:
        raise SystemExit(f"No .xlsx decision tree files found in {source_dir}")

    for workbook_path in workbooks:
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
        workbook_rule_start = len(rules)
        workbook_product_start = len(products)
        scanned_sheets: list[str] = []

        for worksheet in workbook.worksheets:
            if should_skip_sheet(worksheet.title):
                continue
            scanned_sheets.append(worksheet.title)
            sheet_products, sheet_rules = extract_sheet_records(workbook_path, worksheet)
            products.extend(sheet_products)
            rules.extend(sheet_rules)

        workbook_summaries.append(
            {
                "workbook": workbook_path.name,
                "product_line": infer_product_line(workbook_path.name, ""),
                "sheets_scanned": scanned_sheets,
                "product_count": len(products) - workbook_product_start,
                "rule_count": len(rules) - workbook_rule_start,
            }
        )
        workbook.close()

    products = dedupe_products(products)
    rules = assign_rule_ids(dedupe_rules(rules))

    normalized = {
        "generated_from": str(source_dir.relative_to(ROOT)) if source_dir.is_relative_to(ROOT) else str(source_dir),
        "description": (
            "Normalized rule candidates extracted from Decision Tree workbooks. "
            "Rows are deterministic candidates and still keep source cells for review."
        ),
        "source_workbooks": workbook_summaries,
        "product_count": len(products),
        "rule_count": len(rules),
        "counts_by_type": dict(Counter(rule["type"] for rule in rules)),
        "products": products,
        "rules": rules,
    }

    write_json(output_dir / NORMALIZED_JSON.name, normalized)
    write_review_csv(output_dir / REVIEW_CSV.name, rules)

    print(f"workbooks={len(workbooks)}")
    print(f"products={len(products)}")
    print(f"rules={len(rules)}")
    print(f"output_json={output_dir / NORMALIZED_JSON.name}")
    print(f"review_csv={output_dir / REVIEW_CSV.name}")


def extract_sheet_records(workbook_path: Path, worksheet: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    product_line = infer_product_line(workbook_path.name, worksheet.title)
    products: list[dict[str, Any]] = []
    rules: list[dict[str, Any]] = []
    current_step = StepContext()
    scan_columns = min(worksheet.max_column or MAX_SCAN_COLUMNS, MAX_SCAN_COLUMNS)

    for row in worksheet.iter_rows(max_col=scan_columns):
        non_empty_cells = [cell for cell in row if normalize_cell_value(cell.value)]
        if not non_empty_cells:
            continue

        row_text = " | ".join(normalize_cell_value(cell.value) for cell in non_empty_cells)
        product_cell = first_product_id_cell(non_empty_cells)
        step_context = step_context_from_row(non_empty_cells)
        if step_context and not product_cell:
            current_step = step_context
            rule = rule_from_text(
                text=row_text,
                product_line=product_line,
                product_id="",
                step=current_step,
                workbook=workbook_path.name,
                sheet=worksheet.title,
                source_cell=non_empty_cells[0].coordinate,
                source_kind="step_header",
            )
            if rule:
                rules.append(rule)
            continue

        if product_cell:
            product = product_from_row(
                workbook_path=workbook_path,
                worksheet_title=worksheet.title,
                row_cells=non_empty_cells,
                product_cell=product_cell,
                product_line=product_line,
                step=current_step,
            )
            products.append(product)

            rule_text = rule_text_from_product_row(product, row_text)
            rule = rule_from_text(
                text=rule_text,
                product_line=product_line,
                product_id=product["product_id"],
                step=current_step,
                workbook=workbook_path.name,
                sheet=worksheet.title,
                source_cell=product_cell.coordinate,
                source_kind="product_row",
            )
            if rule:
                rules.append(rule)
            continue

        if current_step.step_id and has_rule_signal(row_text):
            rule = rule_from_text(
                text=row_text,
                product_line=product_line,
                product_id="",
                step=current_step,
                workbook=workbook_path.name,
                sheet=worksheet.title,
                source_cell=non_empty_cells[0].coordinate,
                source_kind="note_row",
            )
            if rule:
                rules.append(rule)

    return products, rules


def should_skip_sheet(sheet_name: str) -> bool:
    normalized = sheet_name.casefold().strip()
    return any(token in normalized for token in SKIP_SHEET_TOKENS)


def product_from_row(
    workbook_path: Path,
    worksheet_title: str,
    row_cells: list[Any],
    product_cell: Any,
    product_line: str,
    step: StepContext,
) -> dict[str, Any]:
    product_id = normalize_product_id(product_cell.value)
    cells_after_id = [cell for cell in row_cells if cell.column > product_cell.column]
    text_after_id = [normalize_cell_value(cell.value) for cell in cells_after_id]
    description = first_description(text_after_id)
    numeric_values = [normalize_number(cell.value) for cell in cells_after_id]
    numeric_values = [value for value in numeric_values if value is not None]
    comment = comment_from_cells(text_after_id, description)

    return {
        "product_id": product_id,
        "product_line": product_line,
        "step_id": step.step_id,
        "option_group": step.option_group,
        "short_description": description,
        "list_price": numeric_values[0] if numeric_values else None,
        "bmi_min": numeric_values[1] if len(numeric_values) > 1 else None,
        "comment": comment,
        "source": {
            "workbook": workbook_path.name,
            "sheet": worksheet_title,
            "cell": product_cell.coordinate,
        },
    }


def rule_from_text(
    text: str,
    product_line: str,
    product_id: str,
    step: StepContext,
    workbook: str,
    sheet: str,
    source_cell: str,
    source_kind: str,
) -> dict[str, Any] | None:
    normalized_text = compact_text(text)
    if not has_rule_signal(normalized_text):
        return None

    rule_type = classify_rule_type(normalized_text)
    regions = extract_regions(normalized_text)
    payload = normalized_payload(rule_type, normalized_text, product_id, step, regions)
    strength = "hard_block" if rule_type in {"region_allow", "region_block", "choose_one", "must_select", "require", "exclude"} else "warning" if rule_type == "warning" else "info"

    return {
        "id": "",
        "type": rule_type,
        "effect": strength,
        "review_status": "normalized_candidate",
        "confidence": confidence_for_rule(rule_type, normalized_text),
        "product_line": product_line,
        "product_id": product_id,
        "step_id": step.step_id,
        "option_group": step.option_group,
        "regions": regions,
        "message": normalized_text,
        "payload": payload,
        "source_type": source_kind,
        "source": {
            "workbook": workbook,
            "sheet": sheet,
            "cell": source_cell,
        },
    }


def classify_rule_type(text: str) -> str:
    normalized = text.casefold()
    if re.search(r"\b(pick|choose|select)\s+(exactly\s+)?one\b", normalized):
        return "choose_one"
    if re.search(r"\b(china|canada|us|u\.s\.|usa|united states)\b.*\bonly\b", normalized) or re.search(r"\bonly\b.*\b(china|canada|us|u\.s\.|usa|united states)\b", normalized):
        return "region_allow"
    if re.search(r"\b(except|not for|cannot order in|blocked in|exclude)\b", normalized) and extract_regions(text):
        return "region_block"
    if re.search(r"\b(must select|required selection|need to choose|needs to choose|must choose)\b", normalized):
        return "must_select"
    if re.search(r"\b(cannot|can't|not compatible|not support|not supported|cannot be used with|exclude)\b", normalized):
        return "exclude"
    if re.search(r"\b(requires|required|need to|needs to|not include)\b", normalized):
        return "require"
    if re.search(r"\b(confirm|warning|follow|consult|check with)\b", normalized):
        return "warning"
    return "note"


def normalized_payload(
    rule_type: str,
    text: str,
    product_id: str,
    step: StepContext,
    regions: list[str],
) -> dict[str, Any]:
    if rule_type == "region_allow":
        return {"allowed_regions": regions, "source_text": text}
    if rule_type == "region_block":
        return {"blocked_regions": regions, "source_text": text}
    if rule_type == "choose_one":
        return {"min": 1, "max": 1, "step_id": step.step_id, "option_group": step.option_group, "source_text": text}
    if rule_type == "must_select":
        return {"required_selection_text": text, "step_id": step.step_id, "option_group": step.option_group}
    if rule_type == "exclude":
        return {"cannot_combine_text": text, "product_id": product_id}
    if rule_type == "require":
        return {"when": {"product_id": product_id} if product_id else {"step_id": step.step_id}, "then": {"requirement_text": text}}
    if rule_type == "warning":
        return {"message": text}
    return {"message": text}


def confidence_for_rule(rule_type: str, text: str) -> float:
    if rule_type in {"region_allow", "choose_one"}:
        return 0.9
    if rule_type in {"region_block", "must_select", "exclude", "require"}:
        return 0.75
    if rule_type == "warning":
        return 0.65
    return 0.45


def step_context_from_row(row_cells: list[Any]) -> StepContext | None:
    text_values = [normalize_cell_value(cell.value) for cell in row_cells]
    joined = " | ".join(text_values)
    match = STEP_RE.search(joined)
    if not match:
        return None

    step_id = f"Step {match.group(1)}"
    option_group = ""
    for value in text_values:
        if not value or STEP_RE.search(value) or is_product_id_text(value):
            continue
        if not looks_like_header(value):
            option_group = value
            break
    if not option_group:
        option_group = joined
    return StepContext(step_id=step_id, option_group=compact_text(option_group))


def first_product_id_cell(row_cells: list[Any]) -> Any | None:
    for cell in row_cells:
        if is_product_id_text(normalize_cell_value(cell.value)):
            return cell
    return None


def first_description(values: list[str]) -> str:
    for value in values:
        if not value or looks_like_header(value) or is_product_id_text(value):
            continue
        if normalize_number(value) is not None:
            continue
        return compact_text(value)
    return ""


def comment_from_cells(values: list[str], description: str) -> str:
    comments: list[str] = []
    description_seen = not description
    for value in values:
        if not value or looks_like_header(value) or is_product_id_text(value):
            continue
        if normalize_number(value) is not None:
            continue
        compacted = compact_text(value)
        if not description_seen and compacted == description:
            description_seen = True
            continue
        if description_seen:
            comments.append(compacted)
    return " | ".join(dict.fromkeys(comments))


def rule_text_from_product_row(product: dict[str, Any], row_text: str) -> str:
    comment = str(product.get("comment") or "").strip()
    if comment:
        return comment
    return row_text


def has_rule_signal(text: str) -> bool:
    normalized = text.casefold()
    return any(keyword in normalized for keyword in RULE_KEYWORDS)


def extract_regions(text: str) -> list[str]:
    regions = [region for region, pattern in REGION_PATTERNS.items() if pattern.search(text)]
    return sorted(dict.fromkeys(regions))


def infer_product_line(workbook_name: str, sheet_name: str) -> str:
    combined = f"{workbook_name} {sheet_name}".casefold()
    if "compass" in combined and "fmt" in sheet_name.casefold():
        return "DRX-Compass FMT"
    if "compass" in combined and "otc" in sheet_name.casefold():
        return "DRX-Compass OTC"
    if "compass" in combined:
        return "DRX-Compass OTC/FMT"
    if "rise" in combined:
        return "DRX-Rise"
    if "revolution" in combined:
        return "DRX-Revolution Plus"
    if "evolution" in combined:
        return "DRX-Evolution Plus"
    return Path(workbook_name).stem


def dedupe_products(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str, str]] = set()
    deduped: list[dict[str, Any]] = []
    for product in products:
        source = product.get("source") or {}
        key = (
            str(product.get("product_id") or ""),
            str(product.get("product_line") or ""),
            str(product.get("step_id") or ""),
            str(source.get("sheet") or ""),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(product)
    return deduped


def dedupe_rules(rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str, str, str]] = set()
    deduped: list[dict[str, Any]] = []
    for rule in rules:
        source = rule.get("source") or {}
        key = (
            str(rule.get("type") or ""),
            str(rule.get("product_id") or ""),
            str(rule.get("step_id") or ""),
            str(rule.get("message") or ""),
            str(source.get("workbook") or ""),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(rule)
    return deduped


def assign_rule_ids(rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    assigned: list[dict[str, Any]] = []
    for rule_number, rule in enumerate(rules, start=1):
        rule_with_id = dict(rule)
        rule_with_id["id"] = f"decision_tree:{rule_number:04d}"
        assigned.append(rule_with_id)
    return assigned


def write_review_csv(path: Path, rules: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=REVIEW_COLUMNS)
        writer.writeheader()
        for rule in rules:
            source = rule.get("source") or {}
            row = {
                "review_id": rule["id"].replace("decision_tree", "dt_review"),
                "rule_id": rule["id"],
                "rule_type": rule["type"],
                "current_review_status": rule["review_status"],
                "engine_status": "not_implemented",
                "review_reason": "decision_tree_normalized_candidate",
                "product_line": rule.get("product_line", ""),
                "product_id": rule.get("product_id", ""),
                "step_id": rule.get("step_id", ""),
                "option_group": rule.get("option_group", ""),
                "regions": ";".join(rule.get("regions") or []),
                "strength": rule.get("effect", ""),
                "confidence": str(rule.get("confidence", "")),
                "message": rule.get("message", ""),
                "source_workbook": source.get("workbook", ""),
                "source_sheet": source.get("sheet", ""),
                "source_cell": source.get("cell", ""),
                "normalized_payload": json.dumps(rule.get("payload") or {}, ensure_ascii=False, sort_keys=True),
                "review_decision": "",
                "final_rule_type": rule["type"],
                "final_effect": rule.get("effect", ""),
                "reviewer": "",
                "reviewed_at": "",
                "review_notes": "",
            }
            writer.writerow(row)


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return compact_text(str(value))


def normalize_product_id(value: Any) -> str:
    text = normalize_cell_value(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text.strip()


def normalize_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    text = normalize_cell_value(value).replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def compact_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def is_product_id_text(text: str) -> bool:
    return bool(PRODUCT_ID_RE.match(normalize_product_id(text)))


def looks_like_header(text: str) -> bool:
    normalized = text.casefold().strip()
    return normalized in {
        "cat #",
        "cat#",
        "catalog #",
        "description",
        "short description",
        "comment",
        "comments",
        "list price",
        "list price (usd)",
        "bmi min",
        "former drx evolution cat #",
    }


if __name__ == "__main__":
    main()