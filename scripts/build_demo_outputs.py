from __future__ import annotations

import json
import sys
import textwrap
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_NORMALIZED_RULES = ROOT / "rules" / "decision_tree_normalized_rules.json"
DEFAULT_OUTPUT_DIR = ROOT / "Output Sample" / "Generated Demo"
CLIENT_PDF_NAME = "client_quote_demo.pdf"
AUDIT_EXCEL_NAME = "internal_audit_demo.xlsx"

PAGE_WIDTH = 595
PAGE_HEIGHT = 842
LEFT_MARGIN = 54
RIGHT_MARGIN = 541
TOP_MARGIN = 790
BOTTOM_MARGIN = 58


@dataclass(frozen=True)
class PdfPage:
    commands: list[str]


class SimplePdfDocument:
    def __init__(self) -> None:
        self.pages: list[PdfPage] = []
        self.current_commands: list[str] = []
        self.current_y = TOP_MARGIN

    def heading(self, text: str, size: int = 16) -> None:
        self._ensure_space(28)
        self._text(text, size=size, font="F2")
        self.current_y -= 7

    def paragraph(self, text: str, size: int = 10, max_chars: int = 88) -> None:
        for line in textwrap.wrap(text, width=max_chars) or [""]:
            self._text(line, size=size, font="F1")
        self.current_y -= 5

    def bullet(self, text: str, size: int = 10, max_chars: int = 84) -> None:
        wrapped = textwrap.wrap(text, width=max_chars) or [""]
        for line_number, line in enumerate(wrapped):
            prefix = "- " if line_number == 0 else "  "
            self._text(prefix + line, size=size, font="F1")
        self.current_y -= 2

    def rule(self) -> None:
        self._ensure_space(14)
        self.current_commands.append(f"0.74 0.74 0.74 RG {LEFT_MARGIN} {self.current_y} m {RIGHT_MARGIN} {self.current_y} l S")
        self.current_y -= 16

    def save(self, path: Path) -> None:
        self._finish_page()
        path.write_bytes(build_pdf_bytes(self.pages))

    def _text(self, text: str, size: int, font: str) -> None:
        line_height = size + 5
        self._ensure_space(line_height)
        escaped = escape_pdf_text(text)
        self.current_commands.append(
            f"0 0 0 rg BT /{font} {size} Tf {LEFT_MARGIN} {self.current_y} Td ({escaped}) Tj ET"
        )
        self.current_y -= line_height

    def _ensure_space(self, needed_height: int) -> None:
        if self.current_y - needed_height >= BOTTOM_MARGIN:
            return
        self._finish_page()
        self.current_y = TOP_MARGIN

    def _finish_page(self) -> None:
        if not self.current_commands:
            return
        self.pages.append(PdfPage(commands=self.current_commands))
        self.current_commands = []


def main() -> None:
    normalized_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_NORMALIZED_RULES
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT_DIR
    if not normalized_path.exists():
        raise SystemExit(
            f"Normalized rules not found: {normalized_path}. "
            "Run scripts/normalize_decision_tree_rules.py first."
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    data = read_json(normalized_path)
    client_pdf = output_dir / CLIENT_PDF_NAME
    audit_excel = output_dir / AUDIT_EXCEL_NAME

    build_client_pdf(data, client_pdf)
    build_audit_workbook(data, audit_excel)

    print(f"client_pdf={client_pdf}")
    print(f"audit_excel={audit_excel}")
    print(f"rules={len(data.get('rules', []))}")
    print(f"products={len(data.get('products', []))}")


def build_client_pdf(data: dict[str, Any], path: Path) -> None:
    document = SimplePdfDocument()
    products = list(data.get("products") or [])
    rules = list(data.get("rules") or [])
    line_stats = product_line_stats(products, rules)
    rule_counts = Counter(rule.get("type", "unknown") for rule in rules)
    blocking_count = sum(1 for rule in rules if rule.get("effect") == "hard_block")
    warning_count = sum(1 for rule in rules if rule.get("effect") == "warning")

    document.heading("Customer Quote Demo", size=20)
    document.paragraph(f"Generated {date.today().isoformat()} from normalized Decision Tree rules.")
    document.paragraph(
        "This customer-facing package summarizes the configured product families, "
        "visible validation outcomes, and quotation readiness without exposing internal source cells."
    )
    document.rule()

    document.heading("Solution Scope")
    for product_line, stats in line_stats.items():
        document.bullet(
            f"{product_line}: {stats['products']} catalog items and {stats['rules']} normalized rule candidates."
        )

    document.heading("Configuration Preview")
    for product_line, product_group in sample_products_by_line(products).items():
        document.paragraph(product_line, size=11, max_chars=80)
        for product in product_group:
            description = product.get("short_description") or "No description"
            list_price = format_money(product.get("list_price"))
            price_text = f", list price {list_price}" if list_price else ""
            document.bullet(f"{product.get('product_id')}: {description}{price_text}")

    document.heading("Validation Summary")
    document.paragraph(
        f"The normalized rule set contains {len(rules)} candidate rules: "
        f"{blocking_count} blocking or required-selection checks, "
        f"{warning_count} customer-visible warnings, and "
        f"{rule_counts.get('note', 0)} informational notes."
    )
    for rule_type, count in sorted(rule_counts.items()):
        document.bullet(f"{rule_type}: {count}")

    document.heading("Demo Readiness")
    document.paragraph(
        "For this week's demo, sales-facing output can use this PDF as the quote package, "
        "while the internal audit workbook keeps the complete trace from workbook, sheet, cell, payload, and review status."
    )
    document.save(path)


def build_audit_workbook(data: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    rules_sheet = workbook.create_sheet("Rules")
    products_sheet = workbook.create_sheet("Products")
    workflow_sheet = workbook.create_sheet("Workflow")

    fill_summary_sheet(summary_sheet, data)
    fill_rules_sheet(rules_sheet, list(data.get("rules") or []))
    fill_products_sheet(products_sheet, list(data.get("products") or []))
    fill_workflow_sheet(workflow_sheet)

    for worksheet in workbook.worksheets:
        style_sheet(worksheet)

    workbook.save(path)


def fill_summary_sheet(worksheet: Any, data: dict[str, Any]) -> None:
    products = list(data.get("products") or [])
    rules = list(data.get("rules") or [])
    worksheet.append(["Metric", "Value"])
    worksheet.append(["Generated date", date.today().isoformat()])
    worksheet.append(["Source", data.get("generated_from", "")])
    worksheet.append(["Product count", len(products)])
    worksheet.append(["Rule count", len(rules)])
    worksheet.append([])

    worksheet.append(["Rule type", "Count"])
    for rule_type, count in sorted(Counter(rule.get("type", "unknown") for rule in rules).items()):
        worksheet.append([rule_type, count])
    worksheet.append([])

    worksheet.append(["Product line", "Products", "Rules"])
    for product_line, stats in product_line_stats(products, rules).items():
        worksheet.append([product_line, stats["products"], stats["rules"]])
    worksheet.append([])

    worksheet.append(["Workbook", "Product line", "Sheets scanned", "Products", "Rules"])
    for source in data.get("source_workbooks") or []:
        worksheet.append(
            [
                source.get("workbook", ""),
                source.get("product_line", ""),
                ", ".join(source.get("sheets_scanned") or []),
                source.get("product_count", 0),
                source.get("rule_count", 0),
            ]
        )


def fill_rules_sheet(worksheet: Any, rules: list[dict[str, Any]]) -> None:
    headers = [
        "rule_id",
        "type",
        "effect",
        "review_status",
        "confidence",
        "product_line",
        "product_id",
        "step_id",
        "option_group",
        "regions",
        "message",
        "payload",
        "source_workbook",
        "source_sheet",
        "source_cell",
        "review_decision",
        "review_notes",
    ]
    worksheet.append(headers)
    for rule in rules:
        source = rule.get("source") or {}
        worksheet.append(
            [
                rule.get("id", ""),
                rule.get("type", ""),
                rule.get("effect", ""),
                rule.get("review_status", ""),
                rule.get("confidence", ""),
                rule.get("product_line", ""),
                rule.get("product_id", ""),
                rule.get("step_id", ""),
                rule.get("option_group", ""),
                ";".join(rule.get("regions") or []),
                rule.get("message", ""),
                json.dumps(rule.get("payload") or {}, ensure_ascii=False, sort_keys=True),
                source.get("workbook", ""),
                source.get("sheet", ""),
                source.get("cell", ""),
                "",
                "",
            ]
        )
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"


def fill_products_sheet(worksheet: Any, products: list[dict[str, Any]]) -> None:
    headers = [
        "product_id",
        "product_line",
        "step_id",
        "option_group",
        "short_description",
        "list_price",
        "bmi_min",
        "comment",
        "source_workbook",
        "source_sheet",
        "source_cell",
    ]
    worksheet.append(headers)
    for product in products:
        source = product.get("source") or {}
        worksheet.append(
            [
                product.get("product_id", ""),
                product.get("product_line", ""),
                product.get("step_id", ""),
                product.get("option_group", ""),
                product.get("short_description", ""),
                product.get("list_price"),
                product.get("bmi_min"),
                product.get("comment", ""),
                source.get("workbook", ""),
                source.get("sheet", ""),
                source.get("cell", ""),
            ]
        )
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"


def fill_workflow_sheet(worksheet: Any) -> None:
    worksheet.append(["Step", "Owner", "Command or artifact", "Exit criteria"])
    workflow_rows = [
        (
            "Normalize decision tree rules",
            "Automation",
            "python scripts/normalize_decision_tree_rules.py",
            "rules/decision_tree_normalized_rules.json and review CSV are regenerated.",
        ),
        (
            "Review high-risk rules",
            "Product / Sales Ops",
            "rules/decision_tree_rules_needing_review.csv",
            "region, must_select, choose_one, require, and exclude rows have review decisions.",
        ),
        (
            "Merge approved reviewed rules",
            "Automation",
            "python scripts/merge_reviewed_rules.py <reviewed csv>",
            "rules/merged_rules.json contains approved executable rules.",
        ),
        (
            "Build customer and audit outputs",
            "Automation",
            "python scripts/build_demo_outputs.py",
            "Client PDF and internal audit Excel are created under outputs/demo.",
        ),
        (
            "Demo checkpoint",
            "Team",
            "outputs/demo/client_quote_demo.pdf and outputs/demo/internal_audit_demo.xlsx",
            "Customer view hides source cells; audit view keeps full traceability.",
        ),
    ]
    for workflow_row in workflow_rows:
        worksheet.append(list(workflow_row))


def product_line_stats(products: list[dict[str, Any]], rules: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    product_counts = Counter(product.get("product_line") or "Unknown" for product in products)
    rule_counts = Counter(rule.get("product_line") or "Unknown" for rule in rules)
    product_lines = sorted(set(product_counts) | set(rule_counts))
    return {
        product_line: {"products": product_counts[product_line], "rules": rule_counts[product_line]}
        for product_line in product_lines
    }


def sample_products_by_line(products: list[dict[str, Any]], limit: int = 3) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        product_line = product.get("product_line") or "Unknown"
        if len(grouped[product_line]) >= limit:
            continue
        if product.get("product_id") and product.get("short_description"):
            grouped[product_line].append(product)
    return dict(sorted(grouped.items()))


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    thin_side = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin_side)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:200])
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def build_pdf_bytes(pages: list[PdfPage]) -> bytes:
    if not pages:
        pages = [PdfPage(commands=["BT /F1 10 Tf 54 790 Td (No content) Tj ET"])]

    objects: list[bytes] = [b"", b"", b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>", b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>"]
    page_object_ids: list[int] = []

    for page in pages:
        content = "\n".join(page.commands).encode("latin-1", "replace")
        content_object = b"<< /Length " + str(len(content)).encode("ascii") + b" >>\nstream\n" + content + b"\nendstream"
        content_object_id = len(objects) + 1
        objects.append(content_object)

        page_object_id = len(objects) + 1
        page_object_ids.append(page_object_id)
        page_object = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {PAGE_WIDTH} {PAGE_HEIGHT}] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> "
            f"/Contents {content_object_id} 0 R >>"
        ).encode("ascii")
        objects.append(page_object)

    kids = " ".join(f"{page_object_id} 0 R" for page_object_id in page_object_ids)
    objects[0] = b"<< /Type /Catalog /Pages 2 0 R >>"
    objects[1] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_object_ids)} >>".encode("ascii")

    pdf = b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"
    offsets = [0]
    for object_number, content in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf += f"{object_number} 0 obj\n".encode("ascii") + content + b"\nendobj\n"

    xref_start = len(pdf)
    pdf += f"xref\n0 {len(objects) + 1}\n".encode("ascii")
    pdf += b"0000000000 65535 f \n"
    for offset in offsets[1:]:
        pdf += f"{offset:010d} 00000 n \n".encode("ascii")
    pdf += (
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_start}\n%%EOF\n"
    ).encode("ascii")
    return pdf


def escape_pdf_text(text: str) -> str:
    latin_text = str(text).encode("latin-1", "replace").decode("latin-1")
    return latin_text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def format_money(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return str(value)
    return f"${amount:,.0f}"


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()