from __future__ import annotations

import re
from datetime import date
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Callable, Iterable
from xml.sax.saxutils import escape

from app.natural_language import (
    CHEST_EXAMINATION,
    detect_clinical_use_case,
    detect_system_family,
    parse_discount_rate,
)


DISCOUNT_APPROVAL_THRESHOLD = 0.35
DISCOUNT_RATE_PRECISION = 6

_MONEY_QUANTUM = Decimal("0.01")


def _round_money(value: Any) -> float:
    """Round a monetary amount to cents with Decimal half-up semantics.

    Using Decimal avoids binary float artefacts (e.g. 0.145 rounding down)
    while keeping the public float-based line dictionaries unchanged.
    """
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return 0.0
    return float(amount.quantize(_MONEY_QUANTUM, rounding=ROUND_HALF_UP))

# Cached artefacts that must be discarded whenever the quotation changes.
GENERATED_OUTPUT_KEYS = (
    "quotation_excel",
    "approval_excel",
    "customer_pdf",
    "approval_description",
)

APPROVAL_FINGERPRINT_MISMATCH_MESSAGE = (
    "The quotation changed after submission. "
    "Please submit it for approval again."
)

AUTO_APPROVED = "AUTO_APPROVED"
MANAGER_APPROVAL_REQUIRED = "MANAGER_APPROVAL_REQUIRED"
INVALID_QUOTATION = "INVALID"

MANAGER_NOT_SUBMITTED = "NOT_SUBMITTED"
MANAGER_PENDING = "PENDING"
MANAGER_APPROVED = "APPROVED"
MANAGER_REVISION_REQUESTED = "REVISION_REQUESTED"
MANAGER_REJECTED = "REJECTED"

WELCOME_MESSAGE = (
    "Tell me what the customer needs. Include the product, quantity, region, "
    "currency and any accessories. I will prepare the configuration and ask "
    "for the discount if it is missing."
)

DEMO_A_PROMPT = (
    "ABC Hospital in Singapore needs two DRX Compass systems, "
    "two wireless detectors and three-year warranty. "
    "Please prepare the quotation in USD with a 30% discount."
)

DEMO_B_PROMPT = (
    "XYZ Medical Centre in Malaysia needs one DRX Compass system, "
    "one wireless detector and three-year warranty. "
    "Please prepare the quotation in USD with a 40% discount."
)

NUMBER_WORDS = {
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
    "nine": 9,
    "ten": 10,
}

REGION_ALIASES = {
    "Singapore": ("singapore",),
    "Malaysia": ("malaysia",),
    "United States": ("united states", "u.s.", "usa", "us"),
    "Canada": ("canada",),
    "China": ("china", "prc"),
    "Europe": ("europe", "eu", "emea"),
}

CURRENCY_ALIASES = {
    "USD": ("usd", "us dollar", "us dollars"),
    "SGD": ("sgd", "singapore dollar", "singapore dollars"),
    "MYR": ("myr", "malaysian ringgit", "ringgit"),
    "EUR": ("eur", "euro", "euros"),
    "CNY": ("cny", "rmb", "yuan"),
    "CAD": ("cad", "canadian dollar", "canadian dollars"),
}

MAIN_PRODUCT_PRICE_BOOK = (
    (
        re.compile(r"\b(?:drx[- ]?)?compass\b", re.IGNORECASE),
        "DRX-COMPASS",
        "DRX Compass Digital Radiography System",
        100_000.00,
    ),
    (
        re.compile(r"\brevolution\b", re.IGNORECASE),
        "DRX-REVOLUTION",
        "DRX Revolution Mobile Radiography System",
        125_000.00,
    ),
    (
        re.compile(r"\brise\b", re.IGNORECASE),
        "DRX-RISE",
        "DRX Rise Mobile Radiography System",
        90_000.00,
    ),
)

MAIN_PRODUCT_SHORT_NAMES = {
    "DRX-COMPASS": "DRX Compass",
    "DRX-REVOLUTION": "DRX Revolution",
    "DRX-RISE": "DRX Rise",
}

SYSTEM_VARIANT_FAMILIES = ("OTC", "FMT")

CLINICAL_USE_CASE_LABELS = {CHEST_EXAMINATION: "Chest examination"}

# Deterministic clinical package: only accessories that already exist in
# ``ACCESSORY_PRICE_BOOK`` may be recommended.
CHEST_EXAM_DEFAULT_ACCESSORIES = (
    "Wireless Detector",
    "Wall Stand",
    "Grid",
)

# (main product code, system family, clinical use case) -> default accessories.
CLINICAL_CONFIGURATION_PACKAGES = {
    ("DRX-COMPASS", "OTC", CHEST_EXAMINATION): CHEST_EXAM_DEFAULT_ACCESSORIES,
}

ACCESSORY_PRICE_BOOK = {
    "Wireless Detector": {
        "product_code": "DET-WL-01",
        "description": "Wireless Detector",
        "list_unit_price": 15_000.00,
    },
    "Focus Detector": {
        "product_code": "DET-FOCUS-01",
        "description": "Focus Wireless Detector",
        "list_unit_price": 15_000.00,
    },
    "Three-Year Warranty": {
        "product_code": "WAR-3Y-01",
        "description": "Three-Year Extended Warranty",
        "list_unit_price": 8_000.00,
    },
    "Wall Stand": {
        "product_code": "WALL-STD-01",
        "description": "Radiography Wall Stand",
        "list_unit_price": 20_000.00,
    },
    "Patient Table": {
        "product_code": "TABLE-01",
        "description": "Radiography Patient Table",
        "list_unit_price": 25_000.00,
    },
    "Grid": {
        "product_code": "GRID-01",
        "description": "Radiography Grid",
        "list_unit_price": 2_500.00,
    },
}


class QuotationValidationError(ValueError):
    pass


def normalize_configuration(
    conversation_text: str,
    recommendation: dict[str, Any],
) -> dict[str, Any]:
    main_model = recommendation.get("main_model") or {}
    main_description = str(main_model.get("short_description") or "").strip()
    main_product = _normalize_main_product(conversation_text, main_description)
    quantity = _extract_main_quantity(conversation_text)
    accessories = _extract_accessories(conversation_text, quantity)
    system_variant = build_system_variant(
        main_product,
        detect_system_family(conversation_text),
    )
    clinical_use_case = detect_clinical_use_case(conversation_text)

    configuration = {
        "customer_name": _extract_customer_name(conversation_text),
        "region": _extract_region(conversation_text, recommendation),
        "currency": _extract_currency(conversation_text),
        "main_product": main_product,
        "system_variant": system_variant,
        "clinical_use_case": clinical_use_case,
        "quantity": quantity,
        "accessories": accessories,
        "configuration_description": _configuration_description(
            main_product,
            quantity,
            accessories,
            system_variant,
            clinical_use_case,
        ),
        "discount_rate": parse_discount_rate(conversation_text),
    }
    return apply_clinical_configuration_defaults(configuration, conversation_text)


def merge_configuration(
    previous_configuration: dict[str, Any],
    latest_turn: str,
    full_conversation: str,
    recommendation: dict[str, Any],
) -> dict[str, Any]:
    """Combine the full-conversation configuration with the latest correction.

    The full conversation still produces the base configuration, so nothing that
    was already confirmed is lost. Fields that the sales user restates in the
    latest turn override the base values, which is what makes corrections such
    as "Actually change the region to Malaysia" take effect.
    """
    merged: dict[str, Any] = dict(previous_configuration or {})
    base = normalize_configuration(full_conversation, recommendation)
    for key, value in base.items():
        if value not in (None, "", []):
            merged[key] = value

    merged.update(_explicit_turn_fields(latest_turn, recommendation))
    merged["accessories"] = merge_duplicate_accessories(
        merged.get("accessories") or []
    )
    merged = apply_clinical_configuration_defaults(merged, full_conversation)
    return merged


def build_system_variant(main_product: str, system_family: str | None) -> str:
    """Return the configuration variant name such as ``DRX Compass OTC``.

    The variant never introduces a new main product or a new price. It only
    records how the recognised system is mounted.
    """
    if not main_product or not system_family:
        return ""
    if system_family not in SYSTEM_VARIANT_FAMILIES:
        return ""
    code, _, _ = _main_product_price(main_product)
    short_name = MAIN_PRODUCT_SHORT_NAMES.get(code)
    if not short_name:
        return ""
    return f"{short_name} {system_family}"


def system_variant_family(system_variant: str) -> str:
    """Return the family (``OTC``/``FMT``) recorded in ``system_variant``."""
    family = str(system_variant or "").strip().rsplit(" ", 1)[-1].upper()
    return family if family in SYSTEM_VARIANT_FAMILIES else ""


def clinical_use_case_label(clinical_use_case: str | None) -> str:
    return CLINICAL_USE_CASE_LABELS.get(str(clinical_use_case or ""), "")


def variant_product_description(main_product: str, system_variant: str) -> str:
    """Return the quotation description carrying the configuration variant."""
    description = str(main_product or "")
    variant = str(system_variant or "").strip()
    if not description or not variant:
        return description
    code, _, _ = _main_product_price(description)
    short_name = MAIN_PRODUCT_SHORT_NAMES.get(code, "")
    if not short_name or not variant.startswith(short_name):
        return description
    if variant in description:
        return description
    return description.replace(short_name, variant, 1)


def clinical_package_for(configuration: dict[str, Any]) -> tuple[str, ...]:
    """Return the default accessories of the matching clinical package."""
    main_product = str((configuration or {}).get("main_product") or "")
    if not main_product or not is_supported_main_product(main_product):
        return ()
    code, _, _ = _main_product_price(main_product)
    family = system_variant_family(str(configuration.get("system_variant") or ""))
    clinical_use_case = str(configuration.get("clinical_use_case") or "")
    if not family or not clinical_use_case:
        return ()
    return CLINICAL_CONFIGURATION_PACKAGES.get(
        (code, family, clinical_use_case),
        (),
    )


def apply_clinical_configuration_defaults(
    configuration: dict[str, Any],
    conversation_text: str,
) -> dict[str, Any]:
    """Add the accessories a recognised clinical package expects.

    The rule order is: explicit user input first, then the clinical default,
    then the ordinary defaults. Accessories the user excluded are never added
    and are removed if an earlier turn had picked them up.
    """
    configuration = dict(configuration or {})
    text = str(conversation_text or "")
    accessories = [
        accessory
        for accessory in merge_duplicate_accessories(
            configuration.get("accessories") or []
        )
        if not is_accessory_excluded(text, str(accessory.get("name") or ""))
    ]

    quantity = max(1, int(configuration.get("quantity") or 1))
    selected = {accessory["name"] for accessory in accessories}
    for name in clinical_package_for(configuration):
        if name in selected or is_accessory_excluded(text, name):
            continue
        accessories.append({"name": name, "quantity": quantity})

    configuration["accessories"] = accessories
    configuration["configuration_description"] = _configuration_description(
        str(configuration.get("main_product") or ""),
        quantity,
        accessories,
        str(configuration.get("system_variant") or ""),
        configuration.get("clinical_use_case"),
    )
    return configuration


def _explicit_turn_fields(
    latest_turn: str,
    recommendation: dict[str, Any],
) -> dict[str, Any]:
    """Return only the fields the latest turn states explicitly."""
    turn = (latest_turn or "").strip()
    if not turn:
        return {}

    fields: dict[str, Any] = {}
    customer_name = _extract_customer_name(turn)
    if customer_name:
        fields["customer_name"] = customer_name

    region = _extract_region(turn, {})
    if region:
        fields["region"] = region

    currency = _extract_currency(turn)
    if currency:
        fields["currency"] = currency

    main_product = _last_mentioned_main_product(turn)
    if main_product:
        fields["main_product"] = main_product

    quantity = _explicit_main_quantity(turn)
    if quantity is not None:
        fields["quantity"] = quantity

    discount_rate = parse_discount_rate(turn)
    if discount_rate is not None:
        fields["discount_rate"] = discount_rate

    accessories = _extract_accessories(turn, fields.get("quantity", 1))
    if accessories:
        fields["accessories"] = accessories
    return fields


def merge_duplicate_accessories(
    accessories: Iterable[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Collapse accessories that refer to the same catalog item."""
    merged: dict[str, dict[str, Any]] = {}
    for accessory in accessories:
        name = str(accessory.get("name") or "").strip()
        if not name:
            continue
        quantity = max(1, int(accessory.get("quantity") or 1))
        if name in merged:
            merged[name]["quantity"] = max(merged[name]["quantity"], quantity)
        else:
            merged[name] = {"name": name, "quantity": quantity}
    return list(merged.values())


def merge_duplicate_quotation_lines(
    lines: Iterable[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Collapse quotation lines that share the same product code.

    The explicitly requested quantity wins instead of accumulating repeated
    matches, so a product mentioned twice never produces two quotation rows.
    """
    merged: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for line in lines:
        product_code = str(line.get("product_code") or "").strip()
        current = dict(line)
        if product_code not in merged:
            merged[product_code] = current
            order.append(product_code)
            continue
        existing = merged[product_code]
        existing_quantity = _safe_int(existing.get("quantity"))
        current_quantity = _safe_int(current.get("quantity"))
        existing["quantity"] = max(existing_quantity, current_quantity)
    return [merged[code] for code in order]


def _safe_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def missing_configuration_fields(configuration: dict[str, Any]) -> list[str]:
    missing = []
    labels = (
        ("customer_name", "customer name"),
        ("region", "region"),
        ("currency", "currency"),
        ("main_product", "main product"),
    )
    for key, label in labels:
        value = configuration.get(key)
        if not isinstance(value, str) or not value.strip():
            missing.append(label)

    if not configuration.get("configuration_description"):
        missing.append("configuration")
    return missing


def build_quotation_lines(configuration: dict[str, Any]) -> list[dict[str, Any]]:
    missing = missing_configuration_fields(configuration)
    if missing:
        raise QuotationValidationError(
            "Cannot build a quotation until these fields are provided: "
            + ", ".join(missing)
        )

    discount_rate = configuration.get("discount_rate")
    if discount_rate is None:
        raise QuotationValidationError("A discount rate is required.")
    if not isinstance(discount_rate, (int, float)) or not 0 <= discount_rate <= 1:
        raise QuotationValidationError("Discount rate must be between 0% and 100%.")

    main_code, main_description, main_price = _main_product_price(
        str(configuration["main_product"])
    )
    lines = [
        _new_quotation_line(
            product_code=main_code,
            description=variant_product_description(
                main_description,
                str(configuration.get("system_variant") or ""),
            ),
            quantity=configuration.get("quantity", 1),
            list_unit_price=main_price,
            discount_rate=float(discount_rate),
        )
    ]

    for accessory in configuration.get("accessories") or []:
        name = str(accessory.get("name") or "")
        catalog_item = ACCESSORY_PRICE_BOOK.get(name)
        if not catalog_item:
            continue
        lines.append(
            _new_quotation_line(
                product_code=str(catalog_item["product_code"]),
                description=str(catalog_item["description"]),
                quantity=accessory.get("quantity", 1),
                list_unit_price=float(catalog_item["list_unit_price"]),
                discount_rate=float(discount_rate),
            )
        )

    result = recalculate_quotation(merge_duplicate_quotation_lines(lines))
    if result["errors"]:
        raise QuotationValidationError(" ".join(result["errors"]))
    return result["lines"]


def recalculate_quotation(lines: Iterable[dict[str, Any]]) -> dict[str, Any]:
    normalized_lines: list[dict[str, Any]] = []
    errors: list[str] = []
    warnings: list[str] = []

    for index, line in enumerate(lines, start=1):
        product_code = str(line.get("product_code") or "").strip()
        description = str(line.get("description") or "").strip()
        quantity = _coerce_quantity(line.get("quantity"), index, errors)
        list_unit_price = _coerce_money(
            line.get("list_unit_price"),
            "List Unit Price",
            index,
            errors,
        )
        quotation_unit_price = _coerce_money(
            line.get("quotation_unit_price"),
            "Quotation Unit Price",
            index,
            errors,
            allow_zero=True,
        )

        if not product_code:
            errors.append(f"Line {index}: Product Code is required.")
        if not description:
            errors.append(f"Line {index}: Description is required.")
        if quotation_unit_price > list_unit_price and list_unit_price > 0:
            warning = (
                "Quotation price is higher than list price. "
                "Discount rate has been set to 0%."
            )
            if warning not in warnings:
                warnings.append(warning)

        normalized_lines.append(
            {
                "product_code": product_code,
                "description": description,
                "quantity": quantity,
                "list_unit_price": list_unit_price,
                "quotation_unit_price": quotation_unit_price,
                "list_line_total": _round_money(
                    Decimal(quantity) * Decimal(str(list_unit_price))
                ),
                "quotation_line_total": _round_money(
                    Decimal(quantity) * Decimal(str(quotation_unit_price))
                ),
            }
        )

    list_total = _round_money(
        sum(Decimal(str(line["list_line_total"])) for line in normalized_lines)
    )
    quotation_total = _round_money(
        sum(Decimal(str(line["quotation_line_total"])) for line in normalized_lines)
    )
    if list_total <= 0:
        errors.append("List Total must be greater than 0.")

    discount_rate = calculate_discount_rate(list_total, quotation_total)
    if quotation_total > list_total:
        discount_rate = 0.0
    if not 0 <= discount_rate <= 1:
        errors.append("Discount Rate must be between 0% and 100%.")

    approval_status = (
        INVALID_QUOTATION if errors else get_discount_approval_status(discount_rate)
    )
    return {
        "lines": normalized_lines,
        "list_total": list_total,
        "quotation_total": quotation_total,
        "discount_rate": discount_rate,
        "approval_status": approval_status,
        "errors": errors,
        "warnings": warnings,
    }


def calculate_discount_rate(list_total: float, quotation_total: float) -> float:
    if list_total <= 0:
        return 0.0
    # Decimal + rounding to 6 decimals keeps float noise such as
    # 0.35000000000000003 from pushing an exact 35% quotation over the
    # approval threshold.
    try:
        rate = (Decimal(str(list_total)) - Decimal(str(quotation_total))) / Decimal(
            str(list_total)
        )
    except (InvalidOperation, ZeroDivisionError):
        return 0.0
    return float(
        rate.quantize(Decimal(1).scaleb(-DISCOUNT_RATE_PRECISION), rounding=ROUND_HALF_UP)
    )


def get_discount_approval_status(discount_rate: float) -> str:
    rounded_rate = round(float(discount_rate), DISCOUNT_RATE_PRECISION)
    if not 0 <= rounded_rate <= 1:
        raise QuotationValidationError("Discount rate must be between 0% and 100%.")

    # The 35% boundary is included in Sales authority.
    # Only a discount strictly greater than 35% requires manager approval.
    if rounded_rate <= DISCOUNT_APPROVAL_THRESHOLD:
        return AUTO_APPROVED
    return MANAGER_APPROVAL_REQUIRED


def quotation_export_errors(
    configuration: dict[str, Any],
    totals: dict[str, Any],
) -> list[str]:
    """Return the reasons why a quotation must not be exported."""
    errors: list[str] = list(totals.get("errors") or [])
    if not totals.get("lines"):
        errors.append("The quotation has no lines.")
    if float(totals.get("list_total") or 0) <= 0:
        errors.append("List Total must be greater than 0.")
    if float(totals.get("quotation_total") or 0) < 0:
        errors.append("Quotation Total cannot be negative.")
    if not str((configuration or {}).get("customer_name") or "").strip():
        errors.append("A customer name is required.")
    if not str((configuration or {}).get("currency") or "").strip():
        errors.append("A currency is required.")
    return list(dict.fromkeys(errors))


def can_export_quotation(
    configuration: dict[str, Any],
    totals: dict[str, Any],
) -> bool:
    return not quotation_export_errors(configuration, totals)


def is_customer_pdf_available(
    approval_status: str,
    manager_approval_status: str,
) -> bool:
    return approval_status == AUTO_APPROVED or (
        approval_status == MANAGER_APPROVAL_REQUIRED
        and manager_approval_status == MANAGER_APPROVED
    )


def quotation_fingerprint(lines: Iterable[dict[str, Any]]) -> tuple[tuple[Any, ...], ...]:
    return tuple(
        (
            str(line.get("product_code") or ""),
            int(line.get("quantity") or 0),
            round(float(line.get("list_unit_price") or 0), 2),
            round(float(line.get("quotation_unit_price") or 0), 2),
        )
        for line in lines
    )


def manager_status_after_quotation_change(
    previous_lines: Iterable[dict[str, Any]],
    edited_lines: Iterable[dict[str, Any]],
    current_status: str,
) -> str:
    if quotation_fingerprint(previous_lines) != quotation_fingerprint(edited_lines):
        return MANAGER_NOT_SUBMITTED
    return current_status


def clear_generated_outputs(generated_files: dict[str, Any]) -> None:
    """Drop every cached export so no stale file can be downloaded."""
    for key in GENERATED_OUTPUT_KEYS:
        generated_files.pop(key, None)


def can_manager_approve(
    current_lines: Iterable[dict[str, Any]],
    submitted_fingerprint: Any,
) -> bool:
    """Only allow an approval that still matches the submitted quotation."""
    if not submitted_fingerprint:
        return False
    return quotation_fingerprint(current_lines) == submitted_fingerprint


def build_approval_description(
    quotation_id: str,
    configuration: dict[str, Any],
    totals: dict[str, Any],
) -> str:
    currency = str(configuration.get("currency") or "")
    return (
        f"Quotation {quotation_id} for {configuration.get('customer_name', '')} "
        "requires manager approval.\n\n"
        "Customer:\n"
        f"{configuration.get('customer_name', '')}\n\n"
        "Region:\n"
        f"{configuration.get('region', '')}\n\n"
        "Configuration:\n"
        f"{configuration.get('configuration_description', '')}\n\n"
        "Currency:\n"
        f"{currency}\n\n"
        "List total:\n"
        f"{currency} {totals['list_total']:,.2f}\n\n"
        "Quotation total:\n"
        f"{currency} {totals['quotation_total']:,.2f}\n\n"
        "Discount rate:\n"
        f"{totals['discount_rate']:.1%}\n\n"
        "Approval threshold:\n"
        f"{DISCOUNT_APPROVAL_THRESHOLD:.1%}\n\n"
        "Reason for approval:\n"
        "The proposed discount rate exceeds the 35% Sales approval authority.\n\n"
        "Please review the attached quotation and approve, reject or request revision."
    )


def build_quotation_response(
    configuration: dict[str, Any],
    totals: dict[str, Any],
) -> str:
    currency = str(configuration.get("currency") or "")
    opening = (
        "I have prepared the configuration and quotation.\n\n"
        f"The quotation total is {currency} {totals['quotation_total']:,.2f}.\n\n"
    )
    if totals["approval_status"] == AUTO_APPROVED:
        return (
            opening
            + f"The current discount rate is {totals['discount_rate']:.1%}, which is "
            "within the 35% Sales approval authority.\n\n"
            "The quotation has been automatically approved and the customer PDF is ready."
        )
    return (
        opening
        + f"The current discount rate is {totals['discount_rate']:.1%}, which exceeds "
        "the 35% Sales approval authority.\n\n"
        "Manager approval is required before the customer PDF can be generated."
    )


def generate_quotation_excel(
    quotation_id: str,
    configuration: dict[str, Any],
    totals: dict[str, Any],
    approval_status: str,
    *,
    internal_approval: bool = False,
    approval_description: str = "",
    quotation_date: date | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _require_exportable_quotation(configuration, totals)

    workbook = Workbook()
    quotation_sheet = workbook.active
    quotation_sheet.title = "Quotation"
    effective_date = quotation_date or date.today()

    metadata = [
        ("Quotation ID", quotation_id),
        ("Quotation Date", effective_date.isoformat()),
        ("Customer", configuration.get("customer_name", "")),
        ("Region", configuration.get("region", "")),
        ("Currency", configuration.get("currency", "")),
    ]
    if configuration.get("system_variant"):
        metadata.append(("System Variant", str(configuration["system_variant"])))
    clinical_label = clinical_use_case_label(configuration.get("clinical_use_case"))
    if clinical_label:
        metadata.append(("Clinical Use Case", clinical_label))
    metadata.append(("Approval Status", approval_status))
    for row in metadata:
        quotation_sheet.append(row)

    quotation_sheet.append([])
    headers = [
        "Product Code",
        "Description",
        "Quantity",
        "List Unit Price",
        "Quotation Unit Price",
        "List Line Total",
        "Quotation Line Total",
    ]
    quotation_sheet.append(headers)
    header_row = quotation_sheet.max_row

    for line in totals["lines"]:
        quotation_sheet.append(
            [
                line["product_code"],
                line["description"],
                line["quantity"],
                line["list_unit_price"],
                line["quotation_unit_price"],
                line["list_line_total"],
                line["quotation_line_total"],
            ]
        )

    quotation_sheet.append([])
    quotation_sheet.append(["List Total", totals["list_total"]])
    quotation_sheet.append(["Quotation Total", totals["quotation_total"]])
    quotation_sheet.append(["Discount Rate", totals["discount_rate"]])
    quotation_sheet.append(["Approval Status", approval_status])

    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True)
    for cell in quotation_sheet[header_row]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center")

    for row in quotation_sheet.iter_rows(
        min_row=header_row + 1,
        max_row=header_row + len(totals["lines"]),
        min_col=4,
        max_col=7,
    ):
        for cell in row:
            cell.number_format = '#,##0.00'

    discount_row = header_row + len(totals["lines"]) + 4
    quotation_sheet.cell(discount_row, 2).number_format = "0.0%"
    quotation_sheet.freeze_panes = f"A{header_row + 1}"
    quotation_sheet.auto_filter.ref = (
        f"A{header_row}:G{header_row + len(totals['lines'])}"
    )
    _set_column_widths(quotation_sheet, get_column_letter)

    for cell in quotation_sheet["A"]:
        if cell.row <= len(metadata) or cell.row > header_row + len(totals["lines"]):
            cell.font = Font(bold=True)

    if internal_approval:
        summary = workbook.create_sheet("Approval Summary")
        rows = [
            ("Quotation ID", quotation_id),
            ("Customer", configuration.get("customer_name", "")),
            ("Region", configuration.get("region", "")),
            ("Currency", configuration.get("currency", "")),
            ("List Total", totals["list_total"]),
            ("Quotation Total", totals["quotation_total"]),
            ("Discount Rate", totals["discount_rate"]),
            ("Approval Threshold", DISCOUNT_APPROVAL_THRESHOLD),
            ("Approval Status", approval_status),
            ("Approval Description", approval_description),
        ]
        summary.append(["Field", "Value"])
        for row in rows:
            summary.append(row)
        for cell in summary[1]:
            cell.fill = title_fill
            cell.font = title_font
        for row_number in (6, 7):
            summary.cell(row_number, 2).number_format = '#,##0.00'
        for row_number in (8, 9):
            summary.cell(row_number, 2).number_format = "0.0%"
        summary.cell(11, 2).alignment = Alignment(wrap_text=True, vertical="top")
        summary.freeze_panes = "A2"
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 90

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_customer_pdf(
    quotation_id: str,
    configuration: dict[str, Any],
    totals: dict[str, Any],
    *,
    quotation_date: date | None = None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    _require_exportable_quotation(configuration, totals)

    currency = escape(str(configuration.get("currency") or ""))
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        pageCompression=0,
        title=f"Quotation {quotation_id}",
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Right", parent=styles["BodyText"], alignment=TA_RIGHT))
    story = [
        Paragraph("CUSTOMER QUOTATION", styles["Title"]),
        Spacer(1, 5 * mm),
        Paragraph(f"<b>Quotation Number:</b> {escape(quotation_id)}", styles["BodyText"]),
        Paragraph(
            f"<b>Quotation Date:</b> {(quotation_date or date.today()).isoformat()}",
            styles["BodyText"],
        ),
        Paragraph(
            f"<b>Customer:</b> {escape(str(configuration.get('customer_name') or ''))}",
            styles["BodyText"],
        ),
        Paragraph(
            f"<b>Region:</b> {escape(str(configuration.get('region') or ''))}",
            styles["BodyText"],
        ),
        Paragraph(f"<b>Currency:</b> {currency}", styles["BodyText"]),
        Spacer(1, 6 * mm),
    ]

    clinical_label = clinical_use_case_label(configuration.get("clinical_use_case"))
    if clinical_label:
        story.insert(
            len(story) - 1,
            Paragraph(
                f"<b>Configuration:</b> Configured for "
                f"{escape(clinical_label.casefold())}",
                styles["BodyText"],
            ),
        )

    table_data: list[list[Any]] = [
        ["Product Code", "Description", "Quantity", "Unit Price", "Line Total"]
    ]
    for line in totals["lines"]:
        table_data.append(
            [
                escape(str(line["product_code"])),
                escape(str(line["description"])),
                str(line["quantity"]),
                f"{currency} {line['quotation_unit_price']:,.2f}",
                f"{currency} {line['quotation_line_total']:,.2f}",
            ]
        )

    table = Table(
        table_data,
        colWidths=[30 * mm, 66 * mm, 18 * mm, 30 * mm, 32 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2F3")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FC")]),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
                ("TOPPADDING", (0, 0), (-1, 0), 7),
            ]
        )
    )
    story.extend(
        [
            table,
            Spacer(1, 6 * mm),
            Paragraph(
                f"<b>Quotation Total: {currency} {totals['quotation_total']:,.2f}</b>",
                styles["Right"],
            ),
            Spacer(1, 8 * mm),
            Paragraph(
                "This quotation is valid for 30 days from the quotation date.",
                styles["BodyText"],
            ),
            Spacer(1, 4 * mm),
            Paragraph(
                "Thank you for the opportunity to support your radiography requirements.",
                styles["BodyText"],
            ),
        ]
    )
    document.build(story)
    return output.getvalue()


def _require_exportable_quotation(
    configuration: dict[str, Any],
    totals: dict[str, Any],
) -> None:
    errors = quotation_export_errors(configuration, totals)
    if errors:
        raise QuotationValidationError(
            "Cannot export an invalid quotation. " + " ".join(errors)
        )


def _extract_customer_name(text: str) -> str:
    patterns = (
        re.compile(
            r"^\s*([A-Z][A-Za-z0-9&.' -]{2,80}?)\s+in\s+"
            r"[A-Za-z][A-Za-z .'-]+?\s+needs?\b",
            re.IGNORECASE | re.MULTILINE,
        ),
        re.compile(
            r"\bcustomer(?:\s+name)?\s*(?:is|:)\s*"
            r"([A-Za-z0-9&.' -]{2,80})(?:[,.]|\n|$)",
            re.IGNORECASE,
        ),
        re.compile(
            r"^\s*([A-Z][A-Za-z0-9&.' -]{2,80}?)\s+needs?\b",
            re.MULTILINE,
        ),
    )
    for pattern in patterns:
        matches = pattern.findall(text)
        if matches:
            return str(matches[-1]).strip(" .,")
    return ""


def _extract_region(text: str, recommendation: dict[str, Any]) -> str:
    normalized = text.casefold()
    for region, aliases in REGION_ALIASES.items():
        for alias in aliases:
            if _contains_alias(normalized, alias):
                return region

    request = recommendation.get("request") or {}
    parsed_region = str(request.get("region") or "").casefold()
    parsed_region_map = {
        "us": "United States",
        "canada": "Canada",
        "china": "China",
        "eu": "Europe",
    }
    return parsed_region_map.get(parsed_region, "")


def _extract_currency(text: str) -> str:
    normalized = text.casefold()
    matches: list[tuple[int, str]] = []
    for currency, aliases in CURRENCY_ALIASES.items():
        for alias in aliases:
            for match in re.finditer(
                r"(?<![a-z0-9])" + re.escape(alias) + r"(?![a-z0-9])",
                normalized,
            ):
                matches.append((match.start(), currency))
    return max(matches, default=(-1, ""), key=lambda item: item[0])[1]


def _normalize_main_product(text: str, recommendation_description: str) -> str:
    for pattern, _, description, _ in MAIN_PRODUCT_PRICE_BOOK:
        if pattern.search(text) or pattern.search(recommendation_description):
            return description
    return recommendation_description


def _last_mentioned_main_product(text: str) -> str:
    """Return the main product mentioned last in ``text``.

    Using the last mention makes corrections such as
    "Replace DRX Compass with DRX Revolution" resolve to the replacement.
    """
    matches: list[tuple[int, str]] = []
    for pattern, _, description, _ in MAIN_PRODUCT_PRICE_BOOK:
        found = list(pattern.finditer(text))
        if found:
            matches.append((found[-1].start(), description))
    if not matches:
        return ""
    return max(matches, key=lambda item: item[0])[1]


def _extract_main_quantity(text: str) -> int:
    explicit = _explicit_main_quantity(text)
    return explicit if explicit is not None else 1


def _explicit_main_quantity(text: str) -> int | None:
    count_token = r"\d+|" + "|".join(NUMBER_WORDS)
    patterns = (
        re.compile(
            rf"\b({count_token})\s+(?:drx[- ]?compass\s+)?systems?\b",
            re.IGNORECASE,
        ),
        re.compile(
            rf"\b({count_token})\s+"
            rf"(?:(?!per\b|for\b|each\b|every\b)[\w-]+\s+){{0,5}}systems?\b",
            re.IGNORECASE,
        ),
        re.compile(
            rf"\bquantity\s+(?:should\s+be|shall\s+be|must\s+be|is|to|=)\s*"
            rf"({count_token})\b",
            re.IGNORECASE,
        ),
    )
    matches: list[tuple[int, int]] = []
    for pattern in patterns:
        for match in pattern.finditer(text):
            matches.append((match.start(), _number_value(match.group(1))))
    if not matches:
        return None
    return max(matches, key=lambda item: item[0])[1]


ACCESSORY_PATTERNS = (
    (
        "Wireless Detector",
        re.compile(r"\bwireless\s+detectors?\b", re.IGNORECASE),
    ),
    (
        "Focus Detector",
        re.compile(r"\bfocus(?:\s+wireless)?\s+detectors?\b", re.IGNORECASE),
    ),
    (
        "Three-Year Warranty",
        re.compile(
            r"\b(?:three|3)[ -]?year(?:\s+extended)?\s+warranty\b",
            re.IGNORECASE,
        ),
    ),
    ("Wall Stand", re.compile(r"\bwall\s*stands?\b", re.IGNORECASE)),
    (
        "Patient Table",
        re.compile(r"\b(?:patient\s+|radiography\s+)?tables?\b", re.IGNORECASE),
    ),
    ("Grid", re.compile(r"\bgrids?\b", re.IGNORECASE)),
)

# Used only when no branded detector was recognised, so "2 detectors per system"
# still produces a detector line without duplicating the branded detectors.
GENERIC_DETECTOR_RE = re.compile(r"\bdetectors?\b", re.IGNORECASE)
DETECTOR_ACCESSORY_NAMES = ("Wireless Detector", "Focus Detector")

# "per system" repeats the accessory for every main system in the configuration.
PER_SYSTEM_RE = re.compile(
    r"^[\s,]*(?:per|for)\s+(?:(?:each|every)\s+)?(?:main\s+)?system\b",
    re.IGNORECASE,
)
PER_SYSTEM_LOOKAHEAD = 32


def _extract_accessories(text: str, main_quantity: int = 1) -> list[dict[str, Any]]:
    accessories: list[dict[str, Any]] = []
    focus_pattern = dict(ACCESSORY_PATTERNS)["Focus Detector"]
    focus_spans = [match.span() for match in focus_pattern.finditer(text)]
    for name, pattern in ACCESSORY_PATTERNS:
        matches = list(pattern.finditer(text))
        if name == "Wireless Detector":
            # "Focus wireless detector" already matched the Focus pattern.
            matches = [
                match
                for match in matches
                if not _is_inside_any(match.start(), focus_spans)
            ]
        if not matches:
            continue
        accessories.extend(
            _accessory_from_match(name, text, match, main_quantity)
            for match in matches
        )

    detected_names = {accessory["name"] for accessory in accessories}
    if not detected_names & set(DETECTOR_ACCESSORY_NAMES):
        generic_matches = list(GENERIC_DETECTOR_RE.finditer(text))
        if generic_matches:
            accessories.extend(
                _accessory_from_match(
                    "Wireless Detector",
                    text,
                    match,
                    main_quantity,
                )
                for match in generic_matches
            )
    return merge_duplicate_accessories(accessories)


ACCESSORY_PATTERNS_BY_NAME = dict(ACCESSORY_PATTERNS)

# Lightweight negation detection: "without a grid", "no grid", "exclude the
# grid" must stop the clinical package from adding that accessory.
ACCESSORY_NEGATION_MARKERS = (
    r"without|no|not|non|exclude|excluding|excludes|omit|omitting|omits|"
    r"skip|skipping|drop|remove|removing"
)
ACCESSORY_NEGATION_PATTERNS = {
    name: re.compile(
        rf"\b(?:{ACCESSORY_NEGATION_MARKERS})\b"
        rf"(?:\s+(?:the|a|an|any|extra|additional))*\s+" + pattern.pattern,
        re.IGNORECASE,
    )
    for name, pattern in ACCESSORY_PATTERNS
}


def is_accessory_excluded(text: str, name: str) -> bool:
    """Return True when ``text`` explicitly rules the accessory out."""
    pattern = ACCESSORY_PATTERNS_BY_NAME.get(name)
    negation_pattern = ACCESSORY_NEGATION_PATTERNS.get(name)
    if not text or pattern is None or negation_pattern is None:
        return False

    negations = list(negation_pattern.finditer(text))
    if not negations:
        return False

    negation_spans = [match.span() for match in negations]
    last_negation_end = negations[-1].end()
    for match in pattern.finditer(text):
        if _is_inside_any(match.start(), negation_spans):
            continue
        if match.start() >= last_negation_end:
            # The accessory was requested again after the exclusion.
            return False
    return True


def _accessory_from_match(
    name: str,
    text: str,
    match: re.Match[str],
    main_quantity: int,
) -> dict[str, Any]:
    if name == "Three-Year Warranty":
        return {"name": name, "quantity": 1}
    quantity = _quantity_before_match(text, match.start())
    if _is_per_system_accessory(text, match.end()):
        quantity *= max(1, int(main_quantity or 1))
    return {"name": name, "quantity": quantity}


def _is_per_system_accessory(text: str, end: int) -> bool:
    return bool(PER_SYSTEM_RE.match(text[end : end + PER_SYSTEM_LOOKAHEAD]))


def _is_inside_any(position: int, spans: list[tuple[int, int]]) -> bool:
    return any(start <= position < end for start, end in spans)


def _quantity_before_match(text: str, start: int) -> int:
    prefix = text[max(0, start - 24) : start]
    count_token = r"\d+|" + "|".join(NUMBER_WORDS)
    match = re.search(rf"\b({count_token})\s+$", prefix, re.IGNORECASE)
    return _number_value(match.group(1)) if match else 1


def _number_value(value: str) -> int:
    normalized = value.casefold()
    return NUMBER_WORDS.get(normalized, int(value) if value.isdigit() else 1)


def _configuration_description(
    main_product: str,
    quantity: int,
    accessories: list[dict[str, Any]],
    system_variant: str = "",
    clinical_use_case: str | None = None,
) -> str:
    if not main_product:
        return ""
    head = f"{quantity} x {system_variant} system" if system_variant else (
        f"{quantity} x {main_product}"
    )
    label = clinical_use_case_label(clinical_use_case)
    if label:
        head = f"{head} configured for {label.casefold()}"
    parts = [head]
    parts.extend(
        f"{accessory['quantity']} x {accessory['name']}" for accessory in accessories
    )
    return "; ".join(parts)


def is_supported_main_product(main_product: str) -> bool:
    """Return True when the text matches a system in ``MAIN_PRODUCT_PRICE_BOOK``."""
    text = str(main_product or "").strip()
    if not text:
        return False
    return any(pattern.search(text) for pattern, _, _, _ in MAIN_PRODUCT_PRICE_BOOK)


def _main_product_price(main_product: str) -> tuple[str, str, float]:
    for pattern, code, description, price in MAIN_PRODUCT_PRICE_BOOK:
        if pattern.search(main_product):
            return code, description, price
    return "SYSTEM-01", main_product, 100_000.00


def _new_quotation_line(
    *,
    product_code: str,
    description: str,
    quantity: Any,
    list_unit_price: float,
    discount_rate: float,
) -> dict[str, Any]:
    return {
        "product_code": product_code,
        "description": description,
        "quantity": quantity,
        "list_unit_price": list_unit_price,
        "quotation_unit_price": _round_money(
            Decimal(str(list_unit_price)) * (Decimal(1) - Decimal(str(discount_rate)))
        ),
    }


def _coerce_quantity(value: Any, index: int, errors: list[str]) -> int:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        errors.append(f"Line {index}: Quantity must be a whole number of at least 1.")
        return 0
    if not numeric.is_integer() or numeric < 1:
        errors.append(f"Line {index}: Quantity must be a whole number of at least 1.")
        return max(0, int(numeric))
    return int(numeric)


def _coerce_money(
    value: Any,
    label: str,
    index: int,
    errors: list[str],
    *,
    allow_zero: bool = False,
) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        errors.append(f"Line {index}: {label} must be a number.")
        return 0.0
    minimum_is_valid = numeric >= 0 if allow_zero else numeric > 0
    if not minimum_is_valid:
        comparison = "0 or greater" if allow_zero else "greater than 0"
        errors.append(f"Line {index}: {label} must be {comparison}.")
    return _round_money(numeric)


def _contains_alias(text: str, alias: str) -> bool:
    return bool(
        re.search(
            r"(?<![a-z0-9])" + re.escape(alias.casefold()) + r"(?![a-z0-9])",
            text,
        )
    )


def _set_column_widths(worksheet: Any, column_letter: Callable[[int], str]) -> None:
    widths = {
        1: 22,
        2: 48,
        3: 12,
        4: 20,
        5: 22,
        6: 20,
        7: 22,
    }
    for index, width in widths.items():
        worksheet.column_dimensions[column_letter(index)].width = width
