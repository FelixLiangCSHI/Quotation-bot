import sys
import unittest
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.natural_language import parse_discount_rate
from app.quotation import (
    AUTO_APPROVED,
    DEMO_A_PROMPT,
    MANAGER_APPROVAL_REQUIRED,
    MANAGER_APPROVED,
    MANAGER_NOT_SUBMITTED,
    build_quotation_lines,
    generate_customer_pdf,
    generate_quotation_excel,
    get_discount_approval_status,
    is_customer_pdf_available,
    manager_status_after_quotation_change,
    normalize_configuration,
    recalculate_quotation,
)
from app.recommender import QuoteRecommender
from app.serialization import to_jsonable


class QuotationDemoTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.recommendation = to_jsonable(
            QuoteRecommender().recommend_from_text(DEMO_A_PROMPT)
        )
        cls.configuration = normalize_configuration(
            DEMO_A_PROMPT,
            cls.recommendation,
        )
        cls.lines = build_quotation_lines(cls.configuration)
        cls.totals = recalculate_quotation(cls.lines)

    def test_existing_natural_language_configuration_still_works(self):
        self.assertEqual("6704878", self.recommendation["main_model"]["product_id"])
        self.assertEqual(
            "DRX Compass Digital Radiography System",
            self.configuration["main_product"],
        )
        self.assertEqual(2, self.configuration["quantity"])
        self.assertEqual(2, len(self.configuration["accessories"]))

    def test_parse_30_percent_discount(self):
        self.assertEqual(0.30, parse_discount_rate("30%"))
        self.assertEqual(
            0.30,
            parse_discount_rate("Apply a 30 percent discount"),
        )

    def test_parse_35_percent_discount(self):
        self.assertEqual(0.35, parse_discount_rate("Give the customer 35% off"))

    def test_parse_40_percent_discount(self):
        self.assertEqual(0.40, parse_discount_rate("40% discount"))

    def test_discount_below_35_is_auto_approved(self):
        self.assertEqual(AUTO_APPROVED, get_discount_approval_status(0.30))

    def test_discount_equal_35_is_auto_approved(self):
        self.assertEqual(AUTO_APPROVED, get_discount_approval_status(0.35))

    def test_discount_above_35_requires_manager(self):
        self.assertEqual(
            MANAGER_APPROVAL_REQUIRED,
            get_discount_approval_status(0.350001),
        )

    def test_quotation_edit_recalculates_discount(self):
        edited = [dict(line) for line in self.lines]
        edited[0]["quotation_unit_price"] = 60_000

        totals = recalculate_quotation(edited)

        self.assertNotEqual(self.totals["discount_rate"], totals["discount_rate"])
        self.assertEqual(
            MANAGER_APPROVAL_REQUIRED,
            totals["approval_status"],
        )

    def test_quotation_edit_resets_manager_approval(self):
        edited = [dict(line) for line in self.lines]
        edited[0]["quotation_unit_price"] = 60_000

        manager_status = manager_status_after_quotation_change(
            self.lines,
            edited,
            MANAGER_APPROVED,
        )

        self.assertEqual(MANAGER_NOT_SUBMITTED, manager_status)

    def test_pdf_available_for_auto_approved_quote(self):
        self.assertTrue(
            is_customer_pdf_available(AUTO_APPROVED, MANAGER_NOT_SUBMITTED)
        )

    def test_pdf_locked_before_manager_approval(self):
        self.assertFalse(
            is_customer_pdf_available(
                MANAGER_APPROVAL_REQUIRED,
                MANAGER_NOT_SUBMITTED,
            )
        )

    def test_manager_approval_unlocks_pdf(self):
        self.assertTrue(
            is_customer_pdf_available(
                MANAGER_APPROVAL_REQUIRED,
                MANAGER_APPROVED,
            )
        )

    def test_excel_export_contains_discount_rate(self):
        from openpyxl import load_workbook

        excel_data = generate_quotation_excel(
            "Q-TEST-001",
            self.configuration,
            self.totals,
            self.totals["approval_status"],
        )
        workbook = load_workbook(BytesIO(excel_data), data_only=True)
        values = [
            cell.value
            for row in workbook["Quotation"].iter_rows()
            for cell in row
        ]

        self.assertIn("Discount Rate", values)
        self.assertIn(self.totals["discount_rate"], values)

    def test_customer_pdf_excludes_discount_information(self):
        pdf_data = generate_customer_pdf(
            "Q-TEST-001",
            self.configuration,
            self.totals,
        )

        self.assertTrue(pdf_data.startswith(b"%PDF"))
        self.assertNotIn(b"Discount", pdf_data)
        self.assertNotIn(b"Approval", pdf_data)
        self.assertNotIn(b"List Price", pdf_data)

    def test_price_above_list_sets_discount_to_zero(self):
        edited = [dict(line) for line in self.lines]
        edited[0]["quotation_unit_price"] = 110_000

        totals = recalculate_quotation(edited)

        self.assertEqual(0, totals["discount_rate"])
        self.assertIn(
            "Quotation price is higher than list price. "
            "Discount rate has been set to 0%.",
            totals["warnings"],
        )

    def test_invalid_quantity_is_reported(self):
        edited = [dict(line) for line in self.lines]
        edited[0]["quantity"] = 0

        totals = recalculate_quotation(edited)

        self.assertTrue(totals["errors"])
        self.assertEqual("INVALID", totals["approval_status"])


if __name__ == "__main__":
    unittest.main()
