"""Regression tests for the DRX Compass OTC chest examination scenario.

The tests replay the exact wording used in the presentation, including the
free-form opening sentence "I Need a compass OTC fit best chest examination".
They only call the existing core functions, so no Streamlit runtime, network
call or external service is required.
"""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.natural_language import detect_clinical_use_case, detect_system_family
from app.quotation import (
    AUTO_APPROVED,
    MANAGER_APPROVAL_REQUIRED,
    MANAGER_APPROVED,
    MANAGER_NOT_SUBMITTED,
    build_quotation_lines,
    clinical_use_case_label,
    generate_customer_pdf,
    generate_quotation_excel,
    is_accessory_excluded,
    is_customer_pdf_available,
    merge_configuration,
    normalize_configuration,
    recalculate_quotation,
)
from app.recommender import QuoteRecommender
from app.serialization import to_jsonable

import streamlit_app


LEADER_PROMPT = "I Need a compass OTC fit best chest examination"
COMPASS_DESCRIPTION = "DRX Compass Digital Radiography System"

RECOMMENDER = QuoteRecommender()


def _accessory_quantity(configuration: dict, name: str) -> int:
    for accessory in configuration.get("accessories") or []:
        if accessory["name"] == name:
            return int(accessory["quantity"])
    return 0


def _accessory_names(configuration: dict) -> list[str]:
    return [accessory["name"] for accessory in configuration.get("accessories") or []]


def _configure(text: str) -> dict:
    recommendation = to_jsonable(RECOMMENDER.recommend_from_text(text))
    return normalize_configuration(text, recommendation)


def _advance(configuration: dict, turns: list[str], prompt: str) -> dict:
    """Replay one Streamlit conversation turn without a Streamlit runtime."""
    turns.append(prompt)
    conversation_text = "\n".join(turns)
    recommendation = to_jsonable(RECOMMENDER.recommend_from_text(conversation_text))
    merged = merge_configuration(
        configuration,
        prompt,
        conversation_text,
        recommendation,
    )
    merged, blocked = streamlit_app.guard_unsupported_product(merged)
    plan = streamlit_app.plan_next_reply(merged, blocked)
    totals = None
    if plan["ready"]:
        totals = recalculate_quotation(build_quotation_lines(merged))
    return {"configuration": merged, "plan": plan, "totals": totals}


class ClinicalUseCaseDetectionTests(unittest.TestCase):
    def test_chest_wording_is_detected(self) -> None:
        for text in (
            "chest examination",
            "Chest exam for the emergency room",
            "chest x-ray room",
            "chest xray",
            "thorax imaging",
            "thoracic examination",
            "胸片检查",
            "胸部检查",
            "胸部摄影",
        ):
            with self.subTest(text=text):
                self.assertEqual("chest_examination", detect_clinical_use_case(text))

    def test_generic_qualifiers_are_not_clinical_use_cases(self) -> None:
        for text in ("fit best", "best", "optimal", "suitable", "best suited"):
            with self.subTest(text=text):
                self.assertIsNone(detect_clinical_use_case(text))

    def test_system_family_detection(self) -> None:
        self.assertEqual("OTC", detect_system_family("Compass OTC"))
        self.assertEqual("OTC", detect_system_family("Compass overhead"))
        self.assertEqual("OTC", detect_system_family("Compass ceiling-mounted"))
        self.assertEqual("FMT", detect_system_family("Compass floor mounted"))

    def test_clinical_label(self) -> None:
        self.assertEqual("Chest examination", clinical_use_case_label("chest_examination"))
        self.assertEqual("", clinical_use_case_label(None))


class LeaderFirstTurnTests(unittest.TestCase):
    """13.1 - the original free-form sentence from the leader."""

    def setUp(self) -> None:
        self.turns: list[str] = []
        self.result = _advance({}, self.turns, LEADER_PROMPT)
        self.configuration = self.result["configuration"]

    def test_product_variant_and_clinical_use_case(self) -> None:
        self.assertEqual(COMPASS_DESCRIPTION, self.configuration["main_product"])
        self.assertEqual("DRX Compass OTC", self.configuration["system_variant"])
        self.assertEqual("chest_examination", self.configuration["clinical_use_case"])

    def test_chest_package_accessories(self) -> None:
        self.assertEqual(1, _accessory_quantity(self.configuration, "Wireless Detector"))
        self.assertEqual(1, _accessory_quantity(self.configuration, "Wall Stand"))
        self.assertEqual(1, _accessory_quantity(self.configuration, "Grid"))

    def test_quotation_is_not_generated_yet(self) -> None:
        self.assertIsNone(self.result["totals"])
        self.assertFalse(self.result["plan"]["ready"])

    def test_next_question_is_the_customer_name(self) -> None:
        self.assertEqual("customer_name", self.result["plan"]["field"])
        self.assertIn(
            streamlit_app.FIELD_QUESTIONS["customer_name"],
            self.result["plan"]["reply"],
        )

    def test_reply_keeps_the_recognised_configuration(self) -> None:
        reply = self.result["plan"]["reply"]
        self.assertIn("DRX Compass OTC", reply)
        self.assertIn("Chest examination", reply)
        self.assertIn("Wireless Detector", reply)
        self.assertIn("Wall Stand", reply)
        self.assertIn("Grid", reply)
        self.assertNotIn("fit best", reply)

    def test_configuration_description_hides_internal_details(self) -> None:
        description = self.configuration["configuration_description"]
        self.assertEqual(
            "1 x DRX Compass OTC system configured for chest examination; "
            "1 x Wireless Detector; 1 x Wall Stand; 1 x Grid",
            description,
        )
        self.assertNotIn("fit best", description)


class LeaderMultiTurnFlowTests(unittest.TestCase):
    """13.2 - one question at a time until the quotation can be produced."""

    def setUp(self) -> None:
        self.turns: list[str] = []
        self.configuration: dict = {}
        self.results = []
        for prompt in (
            LEADER_PROMPT,
            "The customer is ABC Hospital.",
            "Use Singapore.",
            "Use USD.",
            "30%.",
        ):
            result = _advance(self.configuration, self.turns, prompt)
            self.configuration = result["configuration"]
            self.results.append(result)

    def test_one_question_per_turn(self) -> None:
        asked = [result["plan"]["field"] for result in self.results]
        self.assertEqual(
            ["customer_name", "region", "currency", "discount_rate", None],
            asked,
        )

    def test_final_configuration(self) -> None:
        self.assertEqual("ABC Hospital", self.configuration["customer_name"])
        self.assertEqual("Singapore", self.configuration["region"])
        self.assertEqual("USD", self.configuration["currency"])
        self.assertEqual("DRX Compass OTC", self.configuration["system_variant"])
        self.assertEqual("chest_examination", self.configuration["clinical_use_case"])
        self.assertAlmostEqual(0.30, self.configuration["discount_rate"])

    def test_quotation_lines_and_approval(self) -> None:
        totals = self.results[-1]["totals"]
        self.assertIsNotNone(totals)
        self.assertEqual(4, len(totals["lines"]))
        self.assertEqual(
            ["DRX-COMPASS", "DET-WL-01", "WALL-STD-01", "GRID-01"],
            [line["product_code"] for line in totals["lines"]],
        )
        self.assertEqual(
            "DRX Compass OTC Digital Radiography System",
            totals["lines"][0]["description"],
        )
        self.assertAlmostEqual(0.30, totals["discount_rate"])
        self.assertEqual(AUTO_APPROVED, totals["approval_status"])

    def test_excel_and_pdf_exports(self) -> None:
        totals = self.results[-1]["totals"]
        excel = generate_quotation_excel(
            "Q-TEST-OTC",
            self.configuration,
            totals,
            totals["approval_status"],
        )
        self.assertTrue(excel)
        self.assertTrue(
            is_customer_pdf_available(totals["approval_status"], MANAGER_NOT_SUBMITTED)
        )
        pdf = generate_customer_pdf("Q-TEST-OTC", self.configuration, totals)
        self.assertTrue(pdf.startswith(b"%PDF"))


class SingleMessageRequestTests(unittest.TestCase):
    """13.3 - the same requirement provided in one message."""

    def setUp(self) -> None:
        self.prompt = (
            "ABC Hospital in Singapore needs one DRX Compass OTC system "
            "best suited for chest examination. "
            "Prepare the quotation in USD with a 30% discount."
        )
        self.configuration = _configure(self.prompt)
        self.totals = recalculate_quotation(build_quotation_lines(self.configuration))

    def test_quotation_is_generated_immediately(self) -> None:
        self.assertIsNone(streamlit_app.next_missing_field(self.configuration))
        self.assertEqual(4, len(self.totals["lines"]))
        self.assertEqual(AUTO_APPROVED, self.totals["approval_status"])

    def test_quotation_contains_the_chest_package(self) -> None:
        descriptions = [line["description"] for line in self.totals["lines"]]
        self.assertIn("DRX Compass OTC Digital Radiography System", descriptions)
        self.assertIn("Wireless Detector", descriptions)
        self.assertIn("Radiography Wall Stand", descriptions)
        self.assertIn("Radiography Grid", descriptions)

    def test_excel_reports_the_variant_and_clinical_use(self) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        excel = generate_quotation_excel(
            "Q-TEST-ONESHOT",
            self.configuration,
            self.totals,
            self.totals["approval_status"],
        )
        workbook = load_workbook(BytesIO(excel))
        sheet = workbook["Quotation"]
        metadata = {
            str(row[0]): str(row[1])
            for row in sheet.iter_rows(min_row=1, max_row=8, max_col=2, values_only=True)
            if row[0]
        }
        self.assertEqual("DRX Compass OTC", metadata.get("System Variant"))
        self.assertEqual("Chest examination", metadata.get("Clinical Use Case"))

        cells = [
            str(cell.value)
            for row in sheet.iter_rows(values_only=False)
            for cell in row
            if cell.value is not None
        ]
        self.assertIn("DRX Compass OTC Digital Radiography System", cells)
        self.assertIn("DRX-COMPASS", cells)
        for forbidden in ("Gross Margin", "COGS", "Cost", "Profit"):
            self.assertNotIn(forbidden, cells)

    def test_customer_pdf_hides_internal_information(self) -> None:
        pdf = generate_customer_pdf("Q-TEST-ONESHOT", self.configuration, self.totals)
        self.assertTrue(pdf.startswith(b"%PDF"))
        for forbidden in (
            b"List Price",
            b"Discount Rate",
            b"Approval Threshold",
            b"Gross Margin",
        ):
            self.assertNotIn(forbidden, pdf)


class QuantityScalingTests(unittest.TestCase):
    """13.4 - the clinical package follows the number of systems."""

    def setUp(self) -> None:
        self.configuration = _configure(
            "ABC Hospital needs two DRX Compass OTC systems "
            "for chest examinations, quoted in USD with a 30% discount. "
            "Region is Singapore."
        )

    def test_quantities(self) -> None:
        self.assertEqual(2, self.configuration["quantity"])
        self.assertEqual(2, _accessory_quantity(self.configuration, "Wireless Detector"))
        self.assertEqual(2, _accessory_quantity(self.configuration, "Wall Stand"))
        self.assertEqual(2, _accessory_quantity(self.configuration, "Grid"))


class ExplicitOverrideTests(unittest.TestCase):
    """13.5 - explicit user input wins over the clinical default."""

    def setUp(self) -> None:
        self.configuration = _configure(
            "One Compass OTC system for chest examination "
            "with two wireless detectors and no grid."
        )

    def test_explicit_quantity_is_kept(self) -> None:
        self.assertEqual(2, _accessory_quantity(self.configuration, "Wireless Detector"))

    def test_default_wall_stand_is_still_added(self) -> None:
        self.assertEqual(1, _accessory_quantity(self.configuration, "Wall Stand"))

    def test_excluded_grid_is_not_added(self) -> None:
        self.assertNotIn("Grid", _accessory_names(self.configuration))

    def test_excluded_wall_stand_is_not_added(self) -> None:
        configuration = _configure(
            "One Compass OTC system for chest examination without a wall stand."
        )
        self.assertNotIn("Wall Stand", _accessory_names(configuration))
        self.assertEqual(1, _accessory_quantity(configuration, "Grid"))

    def test_exclusion_wordings(self) -> None:
        for text in ("without grid", "no grid", "exclude the grid"):
            with self.subTest(text=text):
                self.assertTrue(is_accessory_excluded(text, "Grid"))

    def test_accessory_requested_again_after_an_exclusion(self) -> None:
        self.assertFalse(
            is_accessory_excluded("No grid.\nActually add one grid.", "Grid")
        )


class FmtIsolationTests(unittest.TestCase):
    """13.6 - the OTC chest package must not leak into other families."""

    def test_fmt_does_not_get_the_otc_chest_package(self) -> None:
        configuration = _configure("Compass FMT for chest examination")

        self.assertEqual("DRX Compass FMT", configuration["system_variant"])
        self.assertEqual("chest_examination", configuration["clinical_use_case"])
        self.assertEqual([], configuration["accessories"])

    def test_compass_without_clinical_use_case_keeps_current_behaviour(self) -> None:
        configuration = _configure("ABC Hospital needs one DRX Compass OTC system.")

        self.assertEqual("DRX Compass OTC", configuration["system_variant"])
        self.assertIsNone(configuration["clinical_use_case"])
        self.assertEqual([], configuration["accessories"])


class ManagerApprovalTests(unittest.TestCase):
    """The 35% boundary and the export permissions stay unchanged."""

    def test_forty_percent_requires_manager_approval(self) -> None:
        configuration = _configure(
            "ABC Hospital in Singapore needs one DRX Compass OTC system "
            "for chest examination in USD with a 40% discount."
        )
        totals = recalculate_quotation(build_quotation_lines(configuration))

        self.assertEqual(MANAGER_APPROVAL_REQUIRED, totals["approval_status"])
        self.assertFalse(
            is_customer_pdf_available(totals["approval_status"], MANAGER_NOT_SUBMITTED)
        )
        self.assertTrue(
            is_customer_pdf_available(totals["approval_status"], MANAGER_APPROVED)
        )

    def test_thirty_five_percent_stays_automatic(self) -> None:
        configuration = _configure(
            "ABC Hospital in Singapore needs one DRX Compass OTC system "
            "for chest examination in USD with a 35% discount."
        )
        totals = recalculate_quotation(build_quotation_lines(configuration))

        self.assertEqual(AUTO_APPROVED, totals["approval_status"])


if __name__ == "__main__":
    unittest.main()
