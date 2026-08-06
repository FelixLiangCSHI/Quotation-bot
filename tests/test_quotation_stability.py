import sys
import unittest
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.quotation import (
    AUTO_APPROVED,
    DEMO_A_PROMPT,
    DEMO_B_PROMPT,
    GENERATED_OUTPUT_KEYS,
    MANAGER_APPROVAL_REQUIRED,
    MANAGER_APPROVED,
    MANAGER_NOT_SUBMITTED,
    QuotationValidationError,
    build_approval_description,
    build_quotation_lines,
    calculate_discount_rate,
    can_export_quotation,
    can_manager_approve,
    clear_generated_outputs,
    generate_customer_pdf,
    generate_quotation_excel,
    get_discount_approval_status,
    manager_status_after_quotation_change,
    normalize_configuration,
    quotation_export_errors,
    quotation_fingerprint,
    recalculate_quotation,
)
from app.recommender import QuoteRecommender
from app.serialization import to_jsonable


def _prepare(prompt):
    recommendation = to_jsonable(QuoteRecommender().recommend_from_text(prompt))
    configuration = normalize_configuration(prompt, recommendation)
    lines = build_quotation_lines(configuration)
    return configuration, lines, recalculate_quotation(lines)


class DiscountBoundaryTests(unittest.TestCase):
    def test_discount_just_below_35_auto_approved(self):
        self.assertEqual(AUTO_APPROVED, get_discount_approval_status(0.34999999))

    def test_discount_exactly_35_auto_approved(self):
        self.assertEqual(AUTO_APPROVED, get_discount_approval_status(0.35))

    def test_discount_just_above_35_requires_manager(self):
        self.assertEqual(
            MANAGER_APPROVAL_REQUIRED,
            get_discount_approval_status(0.350001),
        )

    def test_rounding_does_not_escalate_exact_35(self):
        # 0.35 * 260000 does not round-trip exactly in binary floating point.
        list_total = 260_000.0
        quotation_total = list_total - list_total * 0.35

        discount_rate = calculate_discount_rate(list_total, quotation_total)

        self.assertEqual(0.35, discount_rate)
        self.assertEqual(AUTO_APPROVED, get_discount_approval_status(discount_rate))

    def test_exact_35_percent_quotation_is_auto_approved(self):
        totals = recalculate_quotation(
            [
                {
                    "product_code": "SYSTEM-01",
                    "description": "System",
                    "quantity": 3,
                    "list_unit_price": 100_000.0,
                    "quotation_unit_price": 65_000.0,
                }
            ]
        )

        self.assertEqual(0.35, totals["discount_rate"])
        self.assertEqual(AUTO_APPROVED, totals["approval_status"])


class GeneratedOutputStateTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.configuration, cls.lines, cls.totals = _prepare(DEMO_A_PROMPT)

    def test_edit_invalidates_generated_files(self):
        generated_files = {key: b"stale" for key in GENERATED_OUTPUT_KEYS}
        generated_files["unrelated"] = "keep"

        clear_generated_outputs(generated_files)

        for key in GENERATED_OUTPUT_KEYS:
            self.assertNotIn(key, generated_files)
        self.assertEqual("keep", generated_files["unrelated"])

    def test_edit_resets_manager_status(self):
        edited = [dict(line) for line in self.lines]
        edited[0]["quantity"] = edited[0]["quantity"] + 1

        self.assertEqual(
            MANAGER_NOT_SUBMITTED,
            manager_status_after_quotation_change(
                self.lines,
                edited,
                MANAGER_APPROVED,
            ),
        )

    def test_approval_fingerprint_matches_current_quote(self):
        submitted = quotation_fingerprint(self.lines)

        self.assertTrue(can_manager_approve(self.lines, submitted))

    def test_changed_quote_cannot_use_previous_approval(self):
        submitted = quotation_fingerprint(self.lines)
        edited = [dict(line) for line in self.lines]
        edited[0]["quotation_unit_price"] = 55_000.0

        self.assertFalse(can_manager_approve(edited, submitted))

    def test_approval_without_submission_is_rejected(self):
        self.assertFalse(can_manager_approve(self.lines, None))


class ExportGuardTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.configuration, cls.lines, cls.totals = _prepare(DEMO_A_PROMPT)
        cls.b_configuration, cls.b_lines, cls.b_totals = _prepare(DEMO_B_PROMPT)

    def test_valid_quotation_can_export(self):
        self.assertTrue(can_export_quotation(self.configuration, self.totals))

    def test_empty_quotation_cannot_export(self):
        totals = recalculate_quotation([])

        self.assertIn("The quotation has no lines.", quotation_export_errors({}, totals))

    def test_missing_customer_and_currency_block_export(self):
        errors = quotation_export_errors(
            {"customer_name": "", "currency": ""},
            self.totals,
        )

        self.assertIn("A customer name is required.", errors)
        self.assertIn("A currency is required.", errors)

    def test_invalid_quote_cannot_export(self):
        edited = [dict(line) for line in self.lines]
        edited[0]["quantity"] = 0
        totals = recalculate_quotation(edited)

        with self.assertRaises(QuotationValidationError):
            generate_quotation_excel(
                "Q-TEST-INVALID",
                self.configuration,
                totals,
                totals["approval_status"],
            )
        with self.assertRaises(QuotationValidationError):
            generate_customer_pdf("Q-TEST-INVALID", self.configuration, totals)

    def test_auto_approved_excel_can_open(self):
        from openpyxl import load_workbook

        excel_data = generate_quotation_excel(
            "Q-TEST-A",
            self.configuration,
            self.totals,
            self.totals["approval_status"],
        )
        workbook = load_workbook(BytesIO(excel_data), data_only=True)

        self.assertEqual(["Quotation"], workbook.sheetnames)

    def test_internal_approval_excel_has_two_sheets(self):
        from openpyxl import load_workbook

        excel_data = generate_quotation_excel(
            "Q-TEST-B",
            self.b_configuration,
            self.b_totals,
            self.b_totals["approval_status"],
            internal_approval=True,
            approval_description=build_approval_description(
                "Q-TEST-B",
                self.b_configuration,
                self.b_totals,
            ),
        )
        workbook = load_workbook(BytesIO(excel_data), data_only=True)

        self.assertEqual(["Quotation", "Approval Summary"], workbook.sheetnames)

    def test_customer_pdf_starts_with_pdf_header(self):
        pdf_data = generate_customer_pdf("Q-TEST-A", self.configuration, self.totals)

        self.assertTrue(pdf_data.startswith(b"%PDF"))

    def test_customer_pdf_excludes_discount(self):
        pdf_data = generate_customer_pdf("Q-TEST-A", self.configuration, self.totals)

        self.assertNotIn(b"Discount", pdf_data)

    def test_customer_pdf_excludes_approval(self):
        pdf_data = generate_customer_pdf("Q-TEST-A", self.configuration, self.totals)

        self.assertNotIn(b"Approval", pdf_data)

    def test_customer_pdf_excludes_list_price(self):
        pdf_data = generate_customer_pdf("Q-TEST-A", self.configuration, self.totals)

        self.assertNotIn(b"List Price", pdf_data)
        self.assertNotIn(b"List Unit Price", pdf_data)


if __name__ == "__main__":
    unittest.main()
